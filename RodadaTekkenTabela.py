# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'RodadaTekkenTabela.ui'
#
# Created by: PyQt5 UI code generator 5.8.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from CampeonatoDAO import CampeonatoDAO
import openpyxl
from openpyxl import Workbook

class Ui_FormPontuacao(object):

    def __init__(self):
        try:
            arquivo = open("id.txt", "r")
            self.id = arquivo.readline().strip("\n")
            arquivo.close()
            print(self.id)
            self.campeonatoSelecionado = CampeonatoDAO().carregarInformacoesCampeonato(self.id)
            self.tabelaGerada = CampeonatoDAO().gerarTabelaPontuacao(self.id, self.campeonatoSelecionado)
            self.nome = "Tabela do campeonato " + self.campeonatoSelecionado.nome
        except Exception as e:
            print(e)
            print(type(e))

    '''
    def atualizarTabela(self):
        self.tabelaGerada = CampeonatoDAO().gerarTabelaPontuacao(self.id)
        self.preencherTabela()'''

    def colorirTabela(self, item, linhaIndex, Campeonato):
        if Campeonato.primeiroNaFinal is True and Campeonato.teraFase2 is True:
            if linhaIndex == 0: #amarelo para o lider que classificará direto pra final. Isto poderá ser alterado futuramente dependendo das configurações do campeonato
                item.setForeground(QtGui.QColor(0, 0,0))
                item.setBackground(QtGui.QColor(255, 255, 0))
            elif linhaIndex > 0 and linhaIndex < Campeonato.quantosNaFase2 + 1: #verde para quem se classifica pra 2ª fase. 17 poderá ser alterado futuramente dependendo das configurações do campeonato
                if linhaIndex % 2 == 1:
                    item.setForeground(QtGui.QColor(0, 0,0))
                    item.setBackground(QtGui.QColor(0, 255, 119))
                else:
                    item.setForeground(QtGui.QColor(0, 0,0))
                    item.setBackground(QtGui.QColor(200, 255, 200))
            elif linhaIndex >= Campeonato.quantosNaFase2 + 1 and linhaIndex <= len(self.tabelaGerada):# vermelho para quem tá fora; len(self.tabelaGerada) - 1 = penúltimo
                if linhaIndex % 2 == 1:
                    item.setForeground(QtGui.QColor(255, 255,255))
                    item.setBackground(QtGui.QColor(255, 0, 0))
                else:
                    item.setForeground(QtGui.QColor(50, 50, 50))
                    item.setBackground(QtGui.QColor(255, 150, 150))
                if linhaIndex == len(self.tabelaGerada) - 1 and Campeonato.ultimoEliminado is True:
                    item.setForeground(QtGui.QColor(255, 255, 255))
                    item.setBackground(QtGui.QColor(0, 0, 0))
        elif Campeonato.primeiroNaFinal is False and Campeonato.teraFase2 is True:
            if linhaIndex < Campeonato.quantosNaFase2:
                if linhaIndex % 2 == 1:
                    item.setForeground(QtGui.QColor(0, 0,0))
                    item.setBackground(QtGui.QColor(0, 255, 119))
                else:
                    item.setForeground(QtGui.QColor(0, 0,0))
                    item.setBackground(QtGui.QColor(200, 255, 200))
            elif linhaIndex >= Campeonato.quantosNaFase2 and linhaIndex <= len(self.tabelaGerada):# vermelho para quem tá fora; len(self.tabelaGerada) - 1 = penúltimo
                if linhaIndex == len(self.tabelaGerada) - 1 and Campeonato.ultimoEliminado is True:
                    item.setForeground(QtGui.QColor(255, 255, 255))
                    item.setBackground(QtGui.QColor(0, 0, 0))
                elif linhaIndex % 2 == 1:
                    item.setForeground(QtGui.QColor(255, 255,255))
                    item.setBackground(QtGui.QColor(255, 0, 0))
                elif linhaIndex % 2 == 0:
                    item.setForeground(QtGui.QColor(50, 50, 50))
                    item.setBackground(QtGui.QColor(255, 150, 150))

        elif Campeonato.primeiroNaFinal is False and Campeonato.teraFase2 is False:
            if linhaIndex == 0:
                item.setForeground(QtGui.QColor(0, 0,0))
                item.setBackground(QtGui.QColor(255, 255, 0))
            elif linhaIndex == 1:
                item.setForeground(QtGui.QColor(0, 0,0))
                item.setBackground(QtGui.QColor(217, 217, 217))
            elif linhaIndex == 2:
                item.setForeground(QtGui.QColor(0, 0,0))
                item.setBackground(QtGui.QColor(153, 51, 51))
            else:
                if linhaIndex == len(self.tabelaGerada) and Campeonato.ultimoEliminado is True:
                    item.setForeground(QtGui.QColor(255, 255, 255))
                    item.setBackground(QtGui.QColor(0, 0, 0))
                else:
                    if linhaIndex % 2 == 1:
                        item.setForeground(QtGui.QColor(0, 0,0))
                        item.setBackground(QtGui.QColor(0, 204, 204))
                    else:
                        item.setForeground(QtGui.QColor(0, 0,0))
                        item.setBackground(QtGui.QColor(0, 179, 179))



    def gerarAviso(self, icone, titulo, msg):
        erro = QtWidgets.QMessageBox()
        erro.setWindowTitle(titulo)
        erro.setIcon(icone) # 0 - sem ícone, 1 - Informação, 2 - Cuidado, 3 - Erro crítico
        erro.setText(msg)
        erro.exec_()

    def exportarTabela(self):
        try:
            arquivo2 = openpyxl.load_workbook("TabelaTekken.xlsx")
            row_excel = 3
            planilha2 = arquivo2.get_sheet_by_name('Planilha Teste') #carregando a planilha

            self.label.setVisible(True)
            self.progressBar.setVisible(True)
            porcentagemPorLinha = 100.00 / int(len(self.tabelaGerada))
            porcentagem = 0.00

            for linha in self.tabelaGerada:
                planilha2.cell(row=row_excel, column=2).value = linha[0]
                planilha2.cell(row=row_excel, column=3).value = str(linha[1])
                planilha2.cell(row=row_excel, column=4).value = linha[2]
                planilha2.cell(row=row_excel, column=5).value = linha[3]
                planilha2.cell(row=row_excel, column=6).value = linha[4]
                planilha2.cell(row=row_excel, column=7).value = linha[5]
                planilha2.cell(row=row_excel, column=8).value = linha[6]
                planilha2.cell(row=row_excel, column=9).value = linha[7]
                planilha2.cell(row=row_excel, column=10).value = linha[8]
                planilha2.cell(row=row_excel, column=11).value = linha[9]
                planilha2.cell(row=row_excel, column=12).value = linha[10]
                planilha2.cell(row=row_excel, column=13).value = linha[11]
                planilha2.cell(row=row_excel, column=14).value = linha[12]
                planilha2.cell(row=row_excel, column=16).value = linha[13]
                planilha2.cell(row=row_excel, column=17).value = linha[14]
                row_excel += 1
                porcentagem += porcentagemPorLinha
                self.progressBar.setValue(porcentagem)

            arquivo2.save("TabelaTekken.xlsx")
            self.gerarAviso(1, "Rodada Tekken - Tabela do campeonato", "O arquivo TabelaTekken.xlsx foi gerado com sucesso\nA tabela foi exportada com sucesso.")
            self.progressBar.setValue(0)
            self.label.setVisible(False)
            self.progressBar.setVisible(False)
        except FileNotFoundError:
            print("O arquivo TabelaTekken.xlsx nao foi achado")
            tabela = Workbook()
            planilha2 = tabela.active
            planilha2.title = "Planilha Teste"
            planilha2.cell(row=2, column=2).value="Nome"
            planilha2.cell(row=2, column=3).value="Pontos"
            planilha2.cell(row=2, column=4).value="Rodadas"
            planilha2.cell(row=2, column=5).value="Vitórias"
            planilha2.cell(row=2, column=6).value="Empates"
            planilha2.cell(row=2, column=7).value="Derrotas"
            planilha2.cell(row=2, column=8).value="RNP+"
            planilha2.cell(row=2, column=9).value="RNG+"
            planilha2.cell(row=2, column=10).value="DKO"
            planilha2.cell(row=2, column=11).value="RN+"
            planilha2.cell(row=2, column=12).value="RNTO+"
            planilha2.cell(row=2, column=13).value="RN-"
            planilha2.cell(row=2, column=14).value="SR"
            planilha2.cell(row=2, column=16).value="P-"
            planilha2.cell(row=2, column=17).value="G-"
            tabela.save("TabelaTekken.xlsx")
            print("Arquivo da tabela criado")
            self.exportarTabela()
        except Exception as e:
            print(type(e))


    def preencherTabela(self):
        if self.campeonatoSelecionado.temRingOut is False:
            header = ["Lutador", "Pontos", "Rodadas", "V", "E", "D", "Perfects", "Greats", "Double KO", "RND+", "Time Out", "RND-", "SR", "P-", "G-"]
            totalColunas = 15
            self.tabela.setColumnCount(totalColunas)
            colunaSaldoRounds = 12
        else:
            header = ["Lutador", "Pontos", "Rodadas", "V", "E", "D", "Perfects", "Greats", "Double KO", "RND+", "Time Out", "Ring Out", "RND-", "SR", "P-", "G-"]
            totalColunas = 16
            self.tabela.setColumnCount(totalColunas)
            colunaSaldoRounds = 13

        index = 0
        for item in header:
            self.tabela.setHorizontalHeaderItem(index, QtWidgets.QTableWidgetItem(item))
            if index > 0:
                self.tabela.resizeColumnToContents(index)
            index += 1
        self.tabela.setColumnWidth(0,150)
        linhaIndex = 0
        for linha in self.tabelaGerada:
            #linha[0] = nome
            self.tabela.setRowCount(linhaIndex + 1)
            for coluna in range(0, totalColunas):
                if coluna == colunaSaldoRounds and linha[colunaSaldoRounds] > 0:
                    item = QtWidgets.QTableWidgetItem("+ " + str(linha[coluna]))
                else:
                    item = QtWidgets.QTableWidgetItem(str(linha[coluna]))

                if coluna > 0:
                    item.setTextAlignment(QtCore.Qt.AlignCenter)

                # CONFIGURAÇÃO DE COR DE FUNDO DO ITEM
                self.colorirTabela(item, linhaIndex, self.campeonatoSelecionado)
                self.tabela.setItem(linhaIndex, coluna, item)
            linhaIndex += 1


    def setupUi(self, Form):
        Form.setObjectName("FormPontuacao")
        Form.resize(1280, 692)

        self.tabela = QtWidgets.QTableWidget(Form)
        self.tabela.setGeometry(QtCore.QRect(120, 40, 1020, 531))
        self.tabela.setObjectName("tabela")
        self.tabela.horizontalHeader().setStretchLastSection(True)
        self.tabela.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)

        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(790, 590, 100, 70))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.exportarTabela)

        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(60, 590, 291, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label_20")
        self.label.setVisible(False)

        self.progressBar = QtWidgets.QProgressBar(Form)
        self.progressBar.setGeometry(QtCore.QRect(60, 620, 400, 23))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.progressBar.setVisible(False)

        '''self.botaoAtualizar = QtWidgets.QPushButton(Form)
        self.botaoAtualizar.setGeometry(QtCore.QRect(630, 590, 101, 71))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.botaoAtualizar.setFont(font)
        self.botaoAtualizar.setObjectName("pushButton")
        self.botaoAtualizar.clicked.connect(self.atualizarTabela)'''

        self.preencherTabela()

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("FormPontuacao", self.nome))
        self.pushButton.setText(_translate("FormPontuacao", "Exportar\n"
"tabela"))
        self.label.setText(_translate("FormPontuacao", "Exportando tabela..."))
        #self.botaoAtualizar.setText(_translate("FormPontuacao", "Atualizar\ntabela"))

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_FormPontuacao()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
